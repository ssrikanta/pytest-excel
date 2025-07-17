
import re
import pandas as pd
from datetime import datetime
from collections import OrderedDict
import pytest
from openpyxl import load_workbook
from openpyxl.styles import Font


_py_ext_re = re.compile(r"\.py$")

_pytest_excel_config = None


def pytest_addoption(parser):
    group = parser.getgroup("terminal reporting")
    group.addoption(
        '--excelreport',
        action="store",
        dest="excelpath",
        metavar="path",
        default=None,
        help="create excel report file at given path.")
    group.addoption(
        '--excel-report',
        action="store",
        dest="excelpath",
        metavar="path",
        default=None,
        help="create excel report file at given path.")
    parser.addini('excel_report_path', 'Path to create excel report file', default=None)



def pytest_configure(config):
    global _pytest_excel_config
    _pytest_excel_config = config
    config_excelpath = config.getoption('excelpath', default=None)
    ini_excelpath = config.getini('excel_report_path') if hasattr(config, 'getini') else None
    excelpath = config_excelpath or ini_excelpath
    if excelpath:
        config._excel = ExcelReporter(excelpath)
        config.pluginmanager.register(config._excel)


def pytest_unconfigure(config):
    excel = getattr(config, '_excel', None)
    if excel:
        del config._excel
        config.pluginmanager.unregister(excel)


def mangle_test_address(address):
    path, possible_open_bracket, params = address.partition('[')
    names = path.split("::")
    try:
        names.remove('()')
    except ValueError:
        pass

    names[0] = names[0].replace("/", '.')
    names[0] = _py_ext_re.sub("", names[0])
    names[-1] += possible_open_bracket + params
    return names


class ExcelReporter(object):
    def __init__(self, excelpath):
        self.results = []
        self.excelpath = datetime.now().strftime(excelpath)
        self.wbook = None

    def append_warning(self, warning_message):
        if self.results:
            fieldnames = list(self.results[0])
            warning_row = {key: '' for key in fieldnames}
            warning_row['result'] = 'WARNING'
            warning_row['message'] = warning_message
            self.append(warning_row)

    def append(self, result):
        self.results.append(result)

    def create_sheet(self, column_heading):
        self.wbook = pd.DataFrame(columns=column_heading)

    def update_worksheet(self):
        self.wbook = pd.concat([self.wbook, pd.DataFrame(self.results)], ignore_index=False)

    def save_excel(self):
        # Save DataFrame to Excel
        self.wbook.to_excel(self.excelpath, index=False)

        # Apply font color and bold to 'result' column and fix description/markers
        try:
            wb = load_workbook(self.excelpath)
            ws = wb.active
            # Find column indices
            header = [cell.value for cell in ws[1]]
            result_col = header.index('result') + 1 if 'result' in header else None
            desc_col = header.index('description') + 1 if 'description' in header else None
            markers_col = header.index('markers') + 1 if 'markers' in header else None
            # Set font color and bold for result values
            color_map = {
                'PASSED': '008000',   # Green
                'FAILED': 'FF0000',   # Red
                'ERROR': 'FF0000',    # Red
                'SKIPPED': '0000FF',  # Blue
                'XFAILED': 'FFA500',  # Orange
                'XPASSED': '800080',  # Purple
                'WARNING': 'FFFF00',  # Yellow
            }
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if result_col:
                    result = row[result_col-1].value
                    color = color_map.get(str(result).upper(), None)
                    if color:
                        row[result_col-1].font = Font(color=color, bold=True)
                # Ensure description is not None
                if desc_col:
                    if row[desc_col-1].value is None:
                        row[desc_col-1].value = ''
                # Ensure markers is not None
                if markers_col:
                    if row[markers_col-1].value is None:
                        row[markers_col-1].value = ''
            wb.save(self.excelpath)
        except Exception:
            pass  # Don't fail test run if styling fails

    def build_result(self, report, status, message):
        result = OrderedDict()
        names = mangle_test_address(report.nodeid)
        result['suite_name'] = names[-2]
        result['test_name'] = names[-1]
        # Always add the test method docstring to the 'description' column
        description = getattr(report, '_excel_test_doc', None)
        if description:
            description = description.strip()
        else:
            description = ''
        result['description'] = description
        result['result'] = status
        result['duration'] = getattr(report, 'duration', 0.0)
        result['timestamp'] = datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
        result['message'] = message
        result['file_name'] = report.location[0]
        result['full_test_name'] = report.nodeid
        # Only include actual pytest markers, not all keywords
        markers = []
        if hasattr(report, '_excel_test_markers'):
            for  v in report._excel_test_markers:

                # Only add marker names, not test names or built-in keywords
                if hasattr(v, 'kwargs'):
                    markers.append(v.kwargs.get('reason', ''))
        result['markers'] = ', '.join(markers) if markers else ''
        result['Links'] = getattr(report, 'extra_link', '')
        self.append(result)


# Standalone pytest hooks
@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()
    doc = ''

    if hasattr(item, 'own_markers') and len(item.own_markers) > 0:
        setattr(report, '_excel_test_markers', item.own_markers)

    if hasattr(item, 'obj') and hasattr(item.obj, '__doc__') and item.obj.__doc__:
        doc = item.obj.__doc__
    setattr(report, '_excel_test_doc', doc)
    extra_link = None
    for hook in getattr(item.config, 'excel_extra_link_hooks', []):
        link = hook(item, call, report)
        if link:
            extra_link = link
            break
    report.extra_link = extra_link

def pytest_runtest_logreport(report):
    global _pytest_excel_config
    excel = getattr(_pytest_excel_config, '_excel', None) if _pytest_excel_config else None
    if excel:
        if report.passed and report.when == "call":
            excel.build_result(report, "PASSED", None)
        elif report.failed and report.when == "call":
            message = getattr(report.longrepr, 'reprcrash', None)
            if message:
                message = message.message
            else:
                message = str(report.longrepr)
            excel.build_result(report, "FAILED", message)
        elif report.failed:
            excel.build_result(report, "ERROR", str(report.longrepr))
        elif report.skipped:
            if hasattr(report, "wasxfail"):
                status = "XFAILED"
                message = f"expected test failure Reason: {report.wasxfail} "
            else:
                status = "SKIPPED"
                _, _, message = report.longrepr
                if message.startswith("Skipped: "):
                    message = message[9:]
            excel.build_result(report, status, message)

def pytest_sessionfinish(session):
    excel = getattr(session.config, '_excel', None)
    if excel and excel.results:
        excel.append_warning('This is a warning added at the end of the report.')
        fieldnames = list(excel.results[0])
        excel.create_sheet(fieldnames)
        excel.update_worksheet()
        excel.save_excel()

def pytest_terminal_summary(terminalreporter):
    excel = getattr(terminalreporter.config, '_excel', None)
    if excel and excel.results:
        terminalreporter.write_sep("-", f"Excel report generated at: {excel.excelpath}")


